import re
from decimal import Decimal
from io import BytesIO

import pytest
from openpyxl import load_workbook
from pydantic import ValidationError

from app.core.config import Settings
from app.core.enums import FinancialPageDisposition
from app.schemas.financial import (
    FinancialClassificationResult,
    FinancialPageClassification,
)
from app.schemas.page import (
    BoundingRegion,
    CanonicalDocument,
    PageMetadata,
    Point,
    Span,
    TableCell,
    TableResult,
    TextBlock,
)
from app.services.financial import (
    CURRENCY_NAME_CONTEXT,
    CURRENCY_SYMBOLS,
    DATE_CONTEXT_RE,
    FINANCIAL_TERMS_ARABIC_RE,
    FINANCIAL_TERMS_CJK_RE,
    IDENTIFIER_CONTEXT_RE,
    MONETARY_CONTEXT_RE,
    FinancialCandidateSelector,
    FinancialExtractionService,
    financial_result_csv,
    financial_result_xlsx,
    selected_page_ranges,
)


def _selector(*, adjacent_pages: int = 0) -> FinancialCandidateSelector:
    return FinancialCandidateSelector(
        financial_labels={"balance_sheet", "financial_table"},
        include_confidence=0.65,
        exclude_confidence=0.9,
        adjacent_pages=adjacent_pages,
    )


def _region(page_number: int) -> BoundingRegion:
    return BoundingRegion(
        page_number=page_number,
        polygon=[Point(x=0, y=0), Point(x=1, y=0), Point(x=1, y=1)],
    )


def test_selected_page_ranges_are_contiguous_and_capped() -> None:
    assert selected_page_ranges([7, 2, 3, 4, 8, 9, 12], 2) == [
        (2, 3),
        (4, 4),
        (7, 8),
        (9, 9),
        (12, 12),
    ]
    assert selected_page_ranges([], 25) == []
    with pytest.raises(ValueError):
        selected_page_ranges([1], 0)


def test_provider_classification_is_recall_first() -> None:
    result = _selector().from_provider(
        document_id="doc-1",
        classifier_id="finance-classifier",
        classifier_version="3",
        source_page_count=3,
        raw={
            "documents": [
                {
                    "docType": "balance_sheet",
                    "confidence": 0.98,
                    "boundingRegions": [{"pageNumber": 1}],
                },
                {
                    "docType": "non_financial",
                    "confidence": 0.97,
                    "boundingRegions": [{"pageNumber": 3}],
                },
            ]
        },
    )

    assert result.selected_pages == [1, 2]
    assert result.pages[0].disposition == FinancialPageDisposition.FINANCIAL
    assert result.pages[1].disposition == FinancialPageDisposition.UNCERTAIN
    assert result.pages[1].review_required
    assert result.pages[2].disposition == FinancialPageDisposition.NON_FINANCIAL


def test_competing_non_financial_label_cannot_suppress_financial_evidence() -> None:
    result = _selector().from_provider(
        document_id="doc-conflict",
        classifier_id="finance-classifier",
        classifier_version="3",
        source_page_count=1,
        raw={
            "documents": [
                {
                    "docType": "balance_sheet",
                    "confidence": 0.98,
                    "boundingRegions": [{"pageNumber": 1}],
                },
                {
                    "docType": "non_financial",
                    "confidence": 0.99,
                    "boundingRegions": [{"pageNumber": 1}],
                },
            ]
        },
    )

    page = result.pages[0]
    assert page.selected
    assert page.disposition == FinancialPageDisposition.FINANCIAL
    assert page.review_required
    assert "conflicting_non_financial_evidence" in page.reasons
    assert [candidate.label for candidate in page.candidates] == [
        "non_financial",
        "balance_sheet",
    ]


def test_adjacent_pages_are_selected_for_continuation_review() -> None:
    result = _selector(adjacent_pages=1).from_provider(
        document_id="doc-adjacent",
        classifier_id="finance-classifier",
        classifier_version="1",
        source_page_count=3,
        raw={
            "documents": [
                {
                    "doc_type": "balance_sheet",
                    "confidence": 0.99,
                    "bounding_regions": [{"page_number": 2}],
                },
                {
                    "doc_type": "non_financial",
                    "confidence": 0.99,
                    "bounding_regions": [
                        {"page_number": 1},
                        {"page_number": 3},
                    ],
                },
            ]
        },
    )

    assert result.selected_pages == [1, 2, 3]
    assert result.pages[0].source == "adjacent_page"
    assert result.pages[2].review_required


def test_financial_result_normalizes_values_and_reports_issues() -> None:
    classification = FinancialClassificationResult(
        document_id="doc-values",
        classifier_id="test",
        classifier_version="1",
        source_page_count=2,
        pages=[
            FinancialPageClassification(
                page_number=1,
                label="balance_sheet",
                confidence=0.99,
                disposition=FinancialPageDisposition.FINANCIAL,
                selected=True,
                source="azure_custom_classifier",
            ),
            FinancialPageClassification(
                page_number=2,
                label="financial_table",
                confidence=0.6,
                disposition=FinancialPageDisposition.UNCERTAIN,
                selected=True,
                review_required=True,
                source="azure_custom_classifier",
            ),
        ],
    )
    document = CanonicalDocument(
        document_id="doc-values",
        filename="financial.pdf",
        status="normalizing",
        pages=[
            PageMetadata(
                page_number=1,
                page_count=2,
                width=8.5,
                height=11,
                unit="inch",
                source_text="Balance sheet total assets",
            ),
            PageMetadata(
                page_number=2,
                page_count=2,
                width=8.5,
                height=11,
                unit="inch",
                source_text="Continuation",
            ),
        ],
        tables=[
            TableResult(
                table_id="t-1",
                row_count=2,
                column_count=2,
                cells=[
                    TableCell(
                        cell_id="label",
                        row_index=0,
                        column_index=0,
                        content="Total assets",
                        bounding_regions=[_region(1)],
                    ),
                    TableCell(
                        cell_id="amount",
                        row_index=0,
                        column_index=1,
                        content="($1,234.50)",
                        bounding_regions=[_region(1)],
                    ),
                    TableCell(
                        cell_id="ambiguous-label",
                        row_index=1,
                        column_index=0,
                        content="Other assets",
                        bounding_regions=[_region(1)],
                    ),
                    TableCell(
                        cell_id="ambiguous",
                        row_index=1,
                        column_index=1,
                        content="12,34",
                        bounding_regions=[_region(1)],
                    ),
                ],
            )
        ],
    )

    result = FinancialExtractionService().build(
        document,
        classification,
        processing_version="finance-1",
    )

    amount = next(cell for cell in result.tables[0].cells if cell.cell_id == "amount")
    assert amount.value.normalized_value == Decimal("-1234.50")
    assert amount.value.currency == "USD"
    assert amount.value.negative
    assert {issue.code for issue in result.validation.issues} == {
        "ambiguous_monetary_format",
        "selected_page_without_financial_content",
    }
    assert result.validation.review_required
    ambiguous = next(
        cell for cell in result.tables[0].cells if cell.cell_id == "ambiguous"
    )
    assert ambiguous.value.normalized_value is None
    assert ambiguous.value.normalization_status == "ambiguous"
    assert ambiguous.review_required
    assert result.tables[0].review_required


def _rmb_thousands_classification() -> FinancialClassificationResult:
    return FinancialClassificationResult(
        document_id="doc-rmb",
        classifier_id="test",
        classifier_version="1",
        source_page_count=1,
        pages=[
            FinancialPageClassification(
                page_number=1,
                label="balance_sheet",
                confidence=0.99,
                disposition=FinancialPageDisposition.FINANCIAL,
                selected=True,
                source="azure_custom_classifier",
            )
        ],
    )


def _rmb_thousands_document(cell_contents: list[str]) -> CanonicalDocument:
    # Each amount needs an adjacent (preceding-column) label carrying a monetary term,
    # or _semantic_type never classifies it as "monetary_amount" in the first place -
    # matching how a real balance-sheet row ("Cash | 543,348") is laid out.
    cells: list[TableCell] = []
    for index, content in enumerate(cell_contents):
        cells.append(
            TableCell(
                cell_id=f"label-{index}",
                row_index=index,
                column_index=0,
                content="餘額",
                bounding_regions=[_region(1)],
            )
        )
        cells.append(
            TableCell(
                cell_id=f"amount-{index}",
                row_index=index,
                column_index=1,
                content=content,
                bounding_regions=[_region(1)],
            )
        )
    return CanonicalDocument(
        document_id="doc-rmb",
        filename="annual-report.pdf",
        status="normalizing",
        pages=[
            PageMetadata(
                page_number=1,
                page_count=1,
                width=8.5,
                height=11,
                unit="inch",
                source_text="資產負債表 單位：人民幣千元",
            )
        ],
        tables=[
            TableResult(
                table_id="t-1",
                row_count=len(cell_contents),
                column_count=2,
                cells=cells,
            )
        ],
    )


def test_infers_comma_grouping_from_uniform_rmb_thousands_table() -> None:
    # Real "already scaled to thousands" RMB tables commonly have every value under
    # 1,000,000 - a single comma group at most - so no cell alone ever proves comma
    # grouping under the strong-evidence tier (which requires 2+ occurrences of the
    # same separator in one cell). With >= 2 such single-comma cells and none
    # conflicting, the weak-evidence fallback should resolve them, not leave every
    # cell ambiguous.
    document = _rmb_thousands_document(["543,348", "80,227", "118,558"])

    result = FinancialExtractionService().build(
        document,
        _rmb_thousands_classification(),
        processing_version="finance-1",
    )

    values = {cell.cell_id: cell.value for cell in result.tables[0].cells}
    assert values["amount-0"].normalized_value == Decimal("543348")
    assert values["amount-1"].normalized_value == Decimal("80227")
    assert values["amount-2"].normalized_value == Decimal("118558")
    for cell_id, value in values.items():
        if cell_id.startswith("amount-"):
            assert not value.ambiguous
            assert not value.requires_normalized_correction
    assert result.validation.issues == []


def test_uniform_grouping_fallback_does_not_fire_with_one_conflicting_cell() -> None:
    # The exact same table as above, plus one invalidly-grouped single-comma cell
    # (e.g. an OCR artifact or a genuinely different locale). That one conflicting
    # cell must disqualify the whole table from the weak-evidence fallback - all
    # cells, including the previously-unambiguous ones, must stay ambiguous. This is
    # the core safety property: the fallback never guesses through a real conflict.
    document = _rmb_thousands_document(["543,348", "80,227", "118,558", "12,34"])

    result = FinancialExtractionService().build(
        document,
        _rmb_thousands_classification(),
        processing_version="finance-1",
    )

    values = {cell.cell_id: cell.value for cell in result.tables[0].cells}
    for cell_id, value in values.items():
        if cell_id.startswith("amount-"):
            assert value.ambiguous
            assert value.normalized_value is None


def test_locale_safe_normalization_never_guesses_single_separator_values() -> None:
    service = FinancialExtractionService()

    eu = service._normalize("1.234,56")
    us = service._normalize("1,234.56")
    comma_only = service._normalize("12,34")
    dot_only = service._normalize("1.234")

    assert eu.normalized_value == Decimal("1234.56")
    assert us.normalized_value == Decimal("1234.56")
    assert not eu.ambiguous and not us.ambiguous
    assert comma_only.normalized_value is None and comma_only.ambiguous
    assert dot_only.normalized_value is None and dot_only.ambiguous


def test_cjk_magnitude_suffixes_parse_but_stay_flagged_for_review() -> None:
    service = FinancialExtractionService()

    yi = service._normalize("1.5億")
    wan = service._normalize("2000萬元")
    prefixed = service._normalize("人民币2000万")
    negative = service._normalize("(1.5億)")

    assert yi.normalized_value == Decimal("150000000")
    assert wan.normalized_value == Decimal("20000000")
    assert prefixed.normalized_value == Decimal("20000000")
    assert negative.normalized_value == Decimal("-150000000")
    assert negative.negative
    # Always review-gated in this first release, even though a value was computed -
    # matches the app's recall-first philosophy for a newer, less-proven parser.
    for result in (yi, wan, prefixed, negative):
        assert result.ambiguous
        assert result.requires_normalized_correction
        assert result.review_reason_code == "cjk_magnitude_suffix"
        assert result.normalization_status == "parsed"


def test_cjk_magnitude_suffix_does_not_misfire_on_idioms_or_compounds() -> None:
    service = FinancialExtractionService()

    idiom = service._normalize("萬事如意")
    compound = service._normalize("2億3000萬")

    # No leading digit at all - never reaches the magnitude branch, falls through to
    # the ordinary malformed/unparsed handling unchanged.
    assert idiom.normalized_value is None
    # Compound "implied addition" forms are explicitly out of scope - must fall
    # through to ambiguous/unparsed exactly as before this feature existed, not
    # silently parse only the first suffix.
    assert compound.normalized_value is None


def test_fullwidth_digits_and_punctuation_normalize_correctly() -> None:
    service = FinancialExtractionService()

    # A single full-width comma group, like its ASCII equivalent, needs table-wide
    # corroborating evidence (decimal_separator) to resolve unambiguously - this is
    # the same locale-safe guard proven in
    # test_locale_safe_normalization_never_guesses_single_separator_values, now also
    # correctly applying after full-width-to-ASCII translation.
    resolved = service._normalize("１２３，４５６", decimal_separator=".")
    unresolved = service._normalize("１２３，４５６")

    assert resolved.normalized_value == Decimal("123456")
    assert not resolved.ambiguous
    assert unresolved.normalized_value is None
    assert unresolved.ambiguous


def test_only_ambiguous_monetary_values_require_normalized_corrections() -> None:
    classification = FinancialClassificationResult(
        document_id="doc-semantic-values",
        classifier_id="test",
        classifier_version="1",
        source_page_count=1,
        pages=[
            FinancialPageClassification(
                page_number=1,
                label="financial_table",
                confidence=0.99,
                disposition=FinancialPageDisposition.FINANCIAL,
                selected=True,
                source="azure_custom_classifier",
            )
        ],
    )
    rows = [
        ("oxygen-label", "Oxygen saturation", "oxygen", "95%-100%", ""),
        ("wbc-label", "White blood cells", "wbc", "7.1", "×10 /L"),
        ("crp-label", "C-reactive protein", "crp", "4.2", "mg/L"),
        (
            "phone-label",
            "Contact phone",
            "phone",
            "400-800-2026 (工作日 09:00-18:00)",
            "",
        ),
        ("account-label", "Account number", "account", "6222 0000 1234 5678", ""),
        ("transaction-label", "Transaction ID", "transaction", "2026-06-02", ""),
        ("date-label", "Date", "date", "2026-06-02", ""),
        ("amount-label", "Amount", "amount-ambiguous", "1.234", "EUR"),
    ]
    cells: list[TableCell] = []
    for row_index, (label_id, label, value_id, value, unit) in enumerate(rows):
        cells.extend(
            [
                TableCell(
                    cell_id=label_id,
                    row_index=row_index,
                    column_index=0,
                    content=label,
                    bounding_regions=[_region(1)],
                ),
                TableCell(
                    cell_id=value_id,
                    row_index=row_index,
                    column_index=1,
                    content=value,
                    bounding_regions=[_region(1)],
                ),
                TableCell(
                    cell_id=f"{value_id}-unit",
                    row_index=row_index,
                    column_index=2,
                    content=unit,
                    bounding_regions=[_region(1)],
                ),
            ]
        )
    document = CanonicalDocument(
        document_id="doc-semantic-values",
        filename="synthetic.pdf",
        status="normalizing",
        pages=[
            PageMetadata(
                page_number=1,
                page_count=1,
                width=8.5,
                height=11,
                unit="inch",
                source_text="Synthetic financial and supporting values",
            )
        ],
        tables=[
            TableResult(
                table_id="semantic-table",
                row_count=len(rows),
                column_count=3,
                cells=cells,
            )
        ],
    )

    result = FinancialExtractionService().build(
        document,
        classification,
        processing_version="semantic-1",
    )
    values = {
        cell.cell_id: cell.value
        for cell in result.tables[0].cells
    }

    assert values["oxygen"].semantic_type == "percentage_range"
    assert values["wbc"].semantic_type == "measurement"
    assert values["crp"].semantic_type == "measurement"
    assert values["phone"].semantic_type == "phone_number"
    assert values["account"].semantic_type == "account_number"
    assert values["transaction"].semantic_type == "identifier"
    assert values["date"].semantic_type == "date_or_time"
    for cell_id in (
        "oxygen",
        "wbc",
        "crp",
        "phone",
        "account",
        "transaction",
        "date",
    ):
        assert values[cell_id].normalized_value is None
        assert not values[cell_id].requires_normalized_correction
        assert not values[cell_id].requires_currency_correction

    amount = values["amount-ambiguous"]
    assert amount.semantic_type == "monetary_amount"
    assert amount.normalized_value is None
    assert amount.requires_normalized_correction
    assert amount.review_reason_code == "ambiguous_monetary_separator"
    assert {
        issue.cell_id
        for issue in result.validation.issues
        if issue.code == "ambiguous_monetary_format"
    } == {"amount-ambiguous"}


def test_footnote_reference_column_is_not_misclassified_as_a_monetary_amount() -> None:
    # Regression: real Chinese annual reports have a Note/附註 reference column next
    # to nearly every line item in the notes section. The row's own monetary-context label
    # (e.g. 應收賬款 / accounts receivable) used to be enough on its own to sweep
    # ANY digit-bearing neighbor cell into monetary_amount classification - including the
    # non-numeric note-reference marker itself - which then failed numeric parsing and got
    # flagged "malformed_monetary_value", and (via the intentional "one bad cell poisons the
    # whole table's separator inference" rule) could also push otherwise well-formed amounts
    # in the same table into needing review. Two real-world shapes are covered: the marker
    # embedded directly in the cell ("附註5"), and a bare digit whose note-ness is only
    # signaled by the column header.
    classification = FinancialClassificationResult(
        document_id="doc-footnote-column",
        classifier_id="test",
        classifier_version="1",
        source_page_count=1,
        pages=[
            FinancialPageClassification(
                page_number=1,
                label="financial_table",
                confidence=0.99,
                disposition=FinancialPageDisposition.FINANCIAL,
                selected=True,
                source="azure_custom_classifier",
            )
        ],
    )
    cells = [
        # Header row.
        TableCell(
            cell_id="h-item", row_index=0, column_index=0, content="Item",
            kind="columnHeader", bounding_regions=[_region(1)],
        ),
        TableCell(
            cell_id="h-note", row_index=0, column_index=1, content="附註",
            kind="columnHeader", bounding_regions=[_region(1)],
        ),
        TableCell(
            cell_id="h-amount", row_index=0, column_index=2, content="Amount",
            kind="columnHeader", bounding_regions=[_region(1)],
        ),
        # Data row 1: note-reference marker embedded directly in the cell.
        TableCell(
            cell_id="label-1", row_index=1, column_index=0, content="應收賬款",
            bounding_regions=[_region(1)],
        ),
        TableCell(
            cell_id="note-in-cell", row_index=1, column_index=1, content="附註5",
            bounding_regions=[_region(1)],
        ),
        TableCell(
            cell_id="amount-1", row_index=1, column_index=2, content="1,234,567",
            bounding_regions=[_region(1)],
        ),
        # Data row 2: bare-digit note reference, note-ness signaled only by the column header.
        TableCell(
            cell_id="label-2", row_index=2, column_index=0, content="應付賬款",
            bounding_regions=[_region(1)],
        ),
        TableCell(
            cell_id="note-bare-digit", row_index=2, column_index=1, content="5",
            bounding_regions=[_region(1)],
        ),
        TableCell(
            cell_id="amount-2", row_index=2, column_index=2, content="2,345,678",
            bounding_regions=[_region(1)],
        ),
    ]
    document = CanonicalDocument(
        document_id="doc-footnote-column",
        filename="synthetic.pdf",
        status="normalizing",
        pages=[
            PageMetadata(
                page_number=1, page_count=1, width=8.5, height=11, unit="inch",
                source_text="Synthetic notes table with footnote references",
            )
        ],
        tables=[
            TableResult(table_id="notes-table", row_count=3, column_count=3, cells=cells)
        ],
    )

    result = FinancialExtractionService().build(
        document, classification, processing_version="footnote-1"
    )
    values = {cell.cell_id: cell.value for cell in result.tables[0].cells}

    for note_cell_id in ("note-in-cell", "note-bare-digit"):
        assert values[note_cell_id].semantic_type != "monetary_amount", note_cell_id
        assert not values[note_cell_id].requires_normalized_correction, note_cell_id

    # The amount cells right next to the note-reference column must still classify and
    # normalize correctly - the fix must not collaterally misclassify them too.
    for amount_cell_id in ("amount-1", "amount-2"):
        assert values[amount_cell_id].semantic_type == "monetary_amount", amount_cell_id
        assert not values[amount_cell_id].requires_normalized_correction, amount_cell_id


def test_numeric_normalization_rejects_malformed_wrappers_and_mixed_currency() -> None:
    service = FinancialExtractionService()

    unmatched_parenthesis = service._normalize("(1,234.56")
    misplaced_percent = service._normalize("12%34")
    currency_symbols = list(CURRENCY_SYMBOLS)
    mixed_currency = service._normalize(
        f"{currency_symbols[0]}{currency_symbols[1]} 1,234.56"
    )

    for value in (unmatched_parenthesis, misplaced_percent, mixed_currency):
        assert value.normalized_value is None
        assert value.ambiguous
        assert value.normalization_status == "ambiguous"


def test_yen_yuan_symbol_uses_page_context_or_remains_ambiguous() -> None:
    service = FinancialExtractionService()

    contextual = service._normalize(
        "\u00a5 18,520.75", decimal_separator=".", currency_context="CNY"
    )
    unknown = service._normalize("\u00a5 18,520.75", decimal_separator=".")

    assert contextual.normalized_value == Decimal("18520.75")
    assert contextual.currency == "CNY"
    assert contextual.currency_source == "page_context"
    assert not contextual.currency_ambiguous
    assert unknown.normalized_value == Decimal("18520.75")
    assert unknown.currency is None
    assert unknown.currency_candidates == ["CNY", "JPY"]
    assert unknown.currency_ambiguous


def test_multi_page_table_is_atomic_across_classification_boundary() -> None:
    classification = FinancialClassificationResult(
        document_id="doc-multipage",
        classifier_id="test",
        classifier_version="1",
        source_page_count=2,
        pages=[
            FinancialPageClassification(
                page_number=1,
                label="financial_table",
                confidence=0.99,
                disposition=FinancialPageDisposition.FINANCIAL,
                selected=True,
                source="azure_custom_classifier",
            ),
            FinancialPageClassification(
                page_number=2,
                label="non_financial",
                confidence=0.99,
                disposition=FinancialPageDisposition.NON_FINANCIAL,
                selected=False,
                source="azure_custom_classifier",
            ),
        ],
    )
    document = CanonicalDocument(
        document_id="doc-multipage",
        filename="multipage.pdf",
        status="normalizing",
        tables=[
            TableResult(
                table_id="t-multipage",
                row_count=2,
                column_count=1,
                cells=[
                    TableCell(
                        cell_id="p1",
                        row_index=0,
                        column_index=0,
                        content="100",
                        bounding_regions=[_region(1)],
                    ),
                    TableCell(
                        cell_id="p2",
                        row_index=1,
                        column_index=0,
                        content="200",
                        bounding_regions=[_region(2)],
                    ),
                ],
            )
        ],
    )

    result = FinancialExtractionService().build(
        document,
        classification,
        processing_version="finance-2",
    )

    table = result.tables[0]
    assert table.complete
    assert table.source_pages == [1, 2]
    assert table.missing_source_pages == []
    assert table.classification_override_pages == [2]
    assert [cell.cell_id for cell in table.cells] == ["p1", "p2"]
    assert table.review_required
    assert result.validation.error_count == 0
    assert result.validation.issues[0].code == "table_crosses_selection_boundary"


def test_provider_table_with_missing_page_cells_is_blocked_as_incomplete() -> None:
    classification = FinancialClassificationResult(
        document_id="doc-missing-table-page",
        classifier_id="test",
        classifier_version="1",
        source_page_count=2,
        pages=[
            FinancialPageClassification(
                page_number=1,
                label="financial_table",
                confidence=0.99,
                disposition=FinancialPageDisposition.FINANCIAL,
                selected=True,
                source="azure_custom_classifier",
            ),
            FinancialPageClassification(
                page_number=2,
                label="financial_table",
                confidence=0.99,
                disposition=FinancialPageDisposition.FINANCIAL,
                selected=True,
                source="azure_custom_classifier",
            ),
        ],
    )
    document = CanonicalDocument(
        document_id="doc-missing-table-page",
        filename="missing-page.pdf",
        status="normalizing",
        tables=[
            TableResult(
                table_id="t-missing-page",
                row_count=2,
                column_count=1,
                bounding_regions=[_region(1), _region(2)],
                cells=[
                    TableCell(
                        cell_id="p1-only",
                        row_index=0,
                        column_index=0,
                        content="100",
                        bounding_regions=[_region(1)],
                    )
                ],
            )
        ],
    )

    result = FinancialExtractionService().build(
        document,
        classification,
        processing_version="finance-2",
    )

    table = result.tables[0]
    assert not table.complete
    assert table.missing_source_pages == [2]
    assert table.review_required
    assert result.validation.error_count == 1
    assert result.validation.issues[0].code == "incomplete_multi_page_table"


def test_layout_measurement_table_is_uncertain_not_definitively_financial() -> None:
    document = CanonicalDocument(
        document_id="doc-medical-layout",
        filename="medical.pdf",
        status="normalizing",
        pages=[
            PageMetadata(
                page_number=1,
                page_count=1,
                width=8.5,
                height=11,
                unit="inch",
                source_text="White blood cells 7.1 ×10 /L C-reactive protein 4.2 mg/L 95%-100%",
            )
        ],
        tables=[
            TableResult(
                table_id="medical-table",
                row_count=1,
                column_count=1,
                cells=[
                    TableCell(
                        cell_id="medical-value",
                        row_index=0,
                        column_index=0,
                        content="7.1",
                        bounding_regions=[_region(1)],
                    )
                ],
            )
        ],
    )

    classification = _selector().from_layout(document=document, source_page_count=1)

    page = classification.pages[0]
    assert page.disposition == FinancialPageDisposition.UNCERTAIN
    assert page.selected
    assert page.review_required
    assert page.reasons == ["layout_measurement_table_requires_review"]
    assert classification.classifier_version == "2.4"


def test_from_layout_selects_chinese_only_financial_table() -> None:
    # \b-wrapped English term matching never fires on CJK text (no word boundaries
    # between adjacent Chinese characters) - a genuine Chinese-only balance sheet must
    # still be classified FINANCIAL via the CJK vocabulary, not left UNCERTAIN.
    document = CanonicalDocument(
        document_id="doc-cjk-table",
        filename="annual-report.pdf",
        status="normalizing",
        pages=[
            PageMetadata(
                page_number=1,
                page_count=1,
                width=8.5,
                height=11,
                unit="inch",
                source_text="資產負債表 單位：人民幣千元",
            )
        ],
        tables=[
            TableResult(
                table_id="balance-sheet",
                row_count=1,
                column_count=1,
                cells=[
                    TableCell(
                        cell_id="total-assets",
                        row_index=0,
                        column_index=0,
                        content="2,543,348",
                        bounding_regions=[_region(1)],
                    )
                ],
            )
        ],
    )

    classification = _selector().from_layout(document=document, source_page_count=1)

    page = classification.pages[0]
    assert page.disposition == FinancialPageDisposition.FINANCIAL
    assert page.selected
    assert not page.review_required
    assert page.reasons == ["layout_table_and_financial_terms"]


def test_has_financial_signal_detects_chinese_narrative() -> None:
    classification = FinancialClassificationResult(
        document_id="doc-cjk-narrative",
        classifier_id="classifier",
        classifier_version="1",
        source_page_count=1,
        pages=[
            FinancialPageClassification(
                page_number=1,
                label="financial_notes",
                confidence=0.96,
                disposition=FinancialPageDisposition.FINANCIAL,
                selected=True,
                source="azure_custom_classifier",
            )
        ],
    )
    # No bare currency symbol/code touching a digit, and no English financial keyword -
    # only the CJK context term (收入, "revenue") should make this a content item.
    text = "本年度收入表現良好，較去年有所增長。"
    document = CanonicalDocument(
        document_id="doc-cjk-narrative",
        filename="notes.pdf",
        status="normalizing",
        pages=[
            PageMetadata(
                page_number=1,
                page_count=1,
                width=8.5,
                height=11,
                unit="inch",
                source_text=text,
            )
        ],
        blocks=[
            TextBlock(
                block_id="revenue-note",
                reading_order=1,
                source_text=text,
                source_language="zh-Hant",
                translated_text=text,
                spans=[Span(offset=0, length=len(text))],
                bounding_regions=[_region(1)],
            )
        ],
    )

    result = FinancialExtractionService().build(
        document,
        classification,
        processing_version="finance-format-1",
    )

    assert len(result.content_items) == 1
    assert result.validation.issues == []


def test_arabic_attached_prefix_matches_via_substring_not_boundary() -> None:
    # Arabic has spaces BETWEEN separate words, but single-letter prefixes (the
    # definite article "ال", conjunctions/prepositions) attach directly to the
    # following word with no space - "الأصول" ("the assets") has no \w/non-\w
    # transition before "أصول", so \b-wrapped matching fails on real Arabic text even
    # though Arabic isn't spaceless like CJK. Confirmed empirically before this fix.
    text = "تشمل الأصول والخصوم وحقوق الملكية"
    assert not re.search(r"\bأصول\b", text)
    assert FINANCIAL_TERMS_ARABIC_RE.search(text)


def test_from_layout_selects_arabic_only_financial_table() -> None:
    document = CanonicalDocument(
        document_id="doc-arabic-table",
        filename="annual-report.pdf",
        status="normalizing",
        pages=[
            PageMetadata(
                page_number=1,
                page_count=1,
                width=8.5,
                height=11,
                unit="inch",
                source_text="الميزانية العمومية بالآلاف من الريال السعودي",
            )
        ],
        tables=[
            TableResult(
                table_id="balance-sheet",
                row_count=1,
                column_count=1,
                cells=[
                    TableCell(
                        cell_id="total-assets",
                        row_index=0,
                        column_index=0,
                        content="2,543,348",
                        bounding_regions=[_region(1)],
                    )
                ],
            )
        ],
    )

    classification = _selector().from_layout(document=document, source_page_count=1)

    page = classification.pages[0]
    assert page.disposition == FinancialPageDisposition.FINANCIAL
    assert page.selected
    assert not page.review_required
    assert page.reasons == ["layout_table_and_financial_terms"]


def test_has_financial_signal_detects_arabic_narrative() -> None:
    classification = FinancialClassificationResult(
        document_id="doc-arabic-narrative",
        classifier_id="classifier",
        classifier_version="1",
        source_page_count=1,
        pages=[
            FinancialPageClassification(
                page_number=1,
                label="financial_notes",
                confidence=0.96,
                disposition=FinancialPageDisposition.FINANCIAL,
                selected=True,
                source="azure_custom_classifier",
            )
        ],
    )
    # No bare currency symbol/code touching a digit, and no English financial keyword -
    # only the Arabic context term (إيرادات, "revenue") should make this a content item.
    text = "ارتفعت إيرادات الشركة هذا العام مقارنة بالعام الماضي."
    document = CanonicalDocument(
        document_id="doc-arabic-narrative",
        filename="notes.pdf",
        status="normalizing",
        pages=[
            PageMetadata(
                page_number=1,
                page_count=1,
                width=8.5,
                height=11,
                unit="inch",
                source_text=text,
            )
        ],
        blocks=[
            TextBlock(
                block_id="revenue-note",
                reading_order=1,
                source_text=text,
                source_language="ar",
                translated_text=text,
                spans=[Span(offset=0, length=len(text))],
                bounding_regions=[_region(1)],
            )
        ],
    )

    result = FinancialExtractionService().build(
        document,
        classification,
        processing_version="finance-format-1",
    )

    assert len(result.content_items) == 1
    assert result.validation.issues == []


def test_expanded_monetary_terms_arabic_cover_new_statement_vocabulary() -> None:
    for term in ("مصروفات", "تكلفة", "فائدة", "ربح", "خسارة", "قرض", "صافي"):
        assert MONETARY_CONTEXT_RE.search(term), f"{term!r} should match MONETARY_CONTEXT_RE"


def test_currency_name_context_recognizes_arabic_currency_names() -> None:
    assert CURRENCY_NAME_CONTEXT["ريال سعودي"] == "SAR"
    assert CURRENCY_NAME_CONTEXT["درهم إماراتي"] == "AED"
    assert CURRENCY_NAME_CONTEXT["دينار كويتي"] == "KWD"
    assert CURRENCY_NAME_CONTEXT["جنيه مصري"] == "EGP"
    assert CURRENCY_NAME_CONTEXT["يورو"] == "EUR"


def test_financial_exports_preserve_precision_and_block_formulas() -> None:
    classification = _selector().from_layout(
        document=CanonicalDocument(
            document_id="doc-export",
            filename="export.pdf",
            status="normalizing",
            pages=[
                PageMetadata(
                    page_number=1,
                    page_count=1,
                    width=8.5,
                    height=11,
                    unit="inch",
                    source_text="Financial total",
                )
            ],
            tables=[
                TableResult(
                    table_id="t-export",
                    row_count=1,
                    column_count=2,
                    cells=[
                        TableCell(
                            cell_id="formula",
                            row_index=0,
                            column_index=0,
                            content="=2+2",
                            bounding_regions=[_region(1)],
                        ),
                        TableCell(
                            cell_id="precise",
                            row_index=0,
                            column_index=1,
                            content="123456789012345.67",
                            bounding_regions=[_region(1)],
                        ),
                    ],
                )
            ],
        ),
        source_page_count=1,
    )
    document = CanonicalDocument(
        document_id="doc-export",
        filename="export.pdf",
        status="normalizing",
        pages=[
            PageMetadata(
                page_number=1,
                page_count=1,
                width=8.5,
                height=11,
                unit="inch",
                source_text="Financial total",
            )
        ],
        tables=[
            TableResult(
                table_id="t-export",
                row_count=1,
                column_count=2,
                cells=[
                    TableCell(
                        cell_id="formula",
                        row_index=0,
                        column_index=0,
                        content="=2+2",
                        bounding_regions=[_region(1)],
                    ),
                    TableCell(
                        cell_id="precise",
                        row_index=0,
                        column_index=1,
                        content="123456789012345.67",
                        bounding_regions=[_region(1)],
                    ),
                ],
            )
        ],
    )
    result = FinancialExtractionService().build(
        document,
        classification,
        processing_version="finance-1",
    )

    csv_output = financial_result_csv(result)
    assert "cell_origin" in csv_output
    assert "table_integrity_status" in csv_output
    assert "'=2+2" in csv_output
    assert "123456789012345.67" in csv_output
    workbook_bytes = financial_result_xlsx(result)
    assert workbook_bytes.startswith(b"PK")
    workbook = load_workbook(BytesIO(workbook_bytes), data_only=False)
    table_sheet = workbook[workbook.sheetnames[1]]
    assert table_sheet["A1"].value == "'=2+2"
    assert table_sheet["B1"].value == "123456789012345.67"


def test_financial_exports_neutralize_normalized_negative_values() -> None:
    classification = FinancialClassificationResult(
        document_id="doc-negative-export",
        classifier_id="test",
        classifier_version="1",
        source_page_count=1,
        pages=[
            FinancialPageClassification(
                page_number=1,
                label="financial_table",
                confidence=0.99,
                disposition=FinancialPageDisposition.FINANCIAL,
                selected=True,
                source="azure_custom_classifier",
            )
        ],
    )
    document = CanonicalDocument(
        document_id="doc-negative-export",
        filename="negative.pdf",
        status="normalizing",
        pages=[
            PageMetadata(
                page_number=1,
                page_count=1,
                width=8.5,
                height=11,
                unit="inch",
                source_text="Financial total",
            )
        ],
        tables=[
            TableResult(
                table_id="t-negative",
                row_count=1,
                column_count=2,
                cells=[
                    TableCell(
                        cell_id="format-evidence",
                        row_index=0,
                        column_index=0,
                        content="1,234.56",
                        bounding_regions=[_region(1)],
                    ),
                    TableCell(
                        cell_id="negative",
                        row_index=0,
                        column_index=1,
                        content="-1234.50",
                        bounding_regions=[_region(1)],
                    ),
                ],
            )
        ],
    )
    result = FinancialExtractionService().build(
        document,
        classification,
        processing_version="finance-2",
    )

    assert "'-1234.50" in financial_result_csv(result)
    workbook = load_workbook(BytesIO(financial_result_xlsx(result)), data_only=False)
    assert workbook[workbook.sheetnames[1]]["B1"].value == "'-1234.50"


def test_selective_mode_requires_classifier_configuration() -> None:
    with pytest.raises(ValidationError):
        Settings(
            financial_extraction_mode="selective",
            financial_classifier_model_id=None,
        )

def test_format_preserving_content_stream_keeps_only_financial_structure() -> None:
    classification = FinancialClassificationResult(
        document_id="doc-format",
        classifier_id="classifier",
        classifier_version="1",
        source_page_count=1,
        pages=[
            FinancialPageClassification(
                page_number=1,
                label="financial_table",
                confidence=0.98,
                disposition=FinancialPageDisposition.FINANCIAL,
                selected=True,
                source="azure_custom_classifier",
            )
        ],
    )
    document = CanonicalDocument(
        document_id="doc-format",
        filename="format.pdf",
        status="normalizing",
        pages=[
            PageMetadata(
                page_number=1,
                page_count=1,
                width=8.5,
                height=11,
                unit="inch",
                source_text="Q4 Results Reporting period: 2025 Revenue increased by 10%.",
            )
        ],
        blocks=[
            TextBlock(
                block_id="heading",
                reading_order=1,
                source_text="Q4 Results",
                source_language="fr-FR",
                translated_text="Q4 Results",
                role="title",
                spans=[Span(offset=0, length=10)],
                bounding_regions=[_region(1)],
            ),
            TextBlock(
                block_id="period",
                reading_order=2,
                source_text="Reporting period: 2025",
                source_language="fr-FR",
                translated_text="Reporting period: 2025",
                spans=[Span(offset=11, length=22)],
                bounding_regions=[_region(1)],
            ),
            TextBlock(
                block_id="narrative",
                reading_order=3,
                source_text="Revenue increased by 10%.",
                source_language="fr-FR",
                translated_text="Revenue increased by 10%.",
                spans=[Span(offset=34, length=25)],
                bounding_regions=[_region(1)],
            ),
            TextBlock(
                block_id="expense",
                reading_order=4,
                source_text="- Operating expenses: €500",
                source_language="fr-FR",
                translated_text="- Operating expenses: €500",
                spans=[Span(offset=60, length=25)],
                bounding_regions=[_region(1)],
            ),
            TextBlock(
                block_id="unrelated",
                reading_order=5,
                source_text="Customer phone: 12345",
                source_language="fr-FR",
                translated_text="Customer phone: 12345",
                spans=[Span(offset=86, length=21)],
                bounding_regions=[_region(1)],
            ),
        ],
        tables=[
            TableResult(
                table_id="financial-grid",
                row_count=1,
                column_count=2,
                bounding_regions=[_region(1)],
                cells=[
                    TableCell(
                        cell_id="label",
                        row_index=0,
                        column_index=0,
                        content="Revenue",
                        source_language="fr-FR",
                        translated_content="Revenue",
                        spans=[Span(offset=108, length=7)],
                        bounding_regions=[_region(1)],
                    ),
                    TableCell(
                        cell_id="value",
                        row_index=0,
                        column_index=1,
                        content="€1.234,56",
                        source_language="fr-FR",
                        translated_content="€1.234,56",
                        spans=[Span(offset=116, length=9)],
                        bounding_regions=[_region(1)],
                    ),
                ],
            )
        ],
    )

    result = FinancialExtractionService().build(
        document,
        classification,
        processing_version="finance-format-1",
    )

    assert result.schema_version == "financial-result-1.4"
    assert [item.item_type for item in result.content_items] == [
        "heading",
        "key_value",
        "paragraph",
        "list_item",
        "table",
    ]
    assert [item.reading_order for item in result.content_items] == [1, 2, 3, 4, 5]
    assert "unrelated" not in {item.item_id for item in result.content_items}
    period = next(item for item in result.content_items if item.item_id == "period")
    assert (period.key, period.value) == ("Reporting period", "2025")
    assert (period.translated_key, period.translated_value) == (
        "Reporting period",
        "2025",
    )
    table_item = result.content_items[-1]
    assert table_item.table_id == "financial-grid"
    assert table_item.source_language == "fr-FR"
    assert result.tables[0].cells[0].translated_text == "Revenue"
    assert result.tables[0].row_count == 1
    assert result.tables[0].column_count == 2


def test_financial_narrative_does_not_require_a_table() -> None:
    classification = FinancialClassificationResult(
        document_id="doc-narrative",
        classifier_id="classifier",
        classifier_version="1",
        source_page_count=1,
        pages=[
            FinancialPageClassification(
                page_number=1,
                label="financial_notes",
                confidence=0.96,
                disposition=FinancialPageDisposition.FINANCIAL,
                selected=True,
                source="azure_custom_classifier",
            )
        ],
    )
    document = CanonicalDocument(
        document_id="doc-narrative",
        filename="notes.pdf",
        status="normalizing",
        pages=[
            PageMetadata(
                page_number=1,
                page_count=1,
                width=8.5,
                height=11,
                unit="inch",
                source_text="Revenue: €500",
            )
        ],
        blocks=[
            TextBlock(
                block_id="revenue-note",
                reading_order=1,
                source_text="Revenue: €500",
                source_language="fr",
                translated_text="Revenue: €500",
                spans=[Span(offset=0, length=13)],
                bounding_regions=[_region(1)],
            )
        ],
    )

    result = FinancialExtractionService().build(
        document,
        classification,
        processing_version="finance-format-1",
    )

    assert [item.item_type for item in result.content_items] == ["key_value"]
    assert result.tables == []
    assert result.validation.issues == []


def test_expanded_monetary_terms_cjk_cover_new_statement_vocabulary() -> None:
    for term in ("费用", "費用", "成本", "利息", "溢利", "亏损", "虧損", "贷款", "股息", "净额"):
        assert MONETARY_CONTEXT_RE.search(term), f"{term!r} should match MONETARY_CONTEXT_RE"


def test_free_user_header_false_positive_is_documented() -> None:
    # Known, accepted tradeoff of substring matching (same design already accepted for
    # 资产/固定资产) - pinned so a future change to this doesn't silently alter
    # classification without a conscious decision. "费用" (expense) is a genuine
    # contiguous substring of "免费用户数量" ("number of free users").
    assert MONETARY_CONTEXT_RE.search("免费用户数量")


def test_expanded_financial_terms_cjk_cover_new_statement_and_governance_vocabulary() -> None:
    for term in (
        "存货", "商誉", "折旧", "摊销", "经营活动", "投资活动", "筹资活动",
        "董事", "联营公司", "关联方", "审计报告", "持续经营业务", "归属",
        "所得税", "减值", "股东",
    ):
        assert FINANCIAL_TERMS_CJK_RE.search(term), f"{term!r} should match FINANCIAL_TERMS_CJK_RE"


def test_currency_name_context_recognizes_hkd_twd_mop_usd_eur_gbp_names() -> None:
    assert CURRENCY_NAME_CONTEXT["港元"] == "HKD"
    assert CURRENCY_NAME_CONTEXT["港幣"] == "HKD"
    assert CURRENCY_NAME_CONTEXT["新台幣"] == "TWD"
    assert CURRENCY_NAME_CONTEXT["澳門元"] == "MOP"
    assert CURRENCY_NAME_CONTEXT["美元"] == "USD"
    assert CURRENCY_NAME_CONTEXT["歐元"] == "EUR"
    assert CURRENCY_NAME_CONTEXT["英鎊"] == "GBP"


def test_page_with_multiple_currency_names_disables_single_currency_context() -> None:
    document = CanonicalDocument(
        document_id="doc-multi-currency",
        filename="annual-report.pdf",
        status="normalizing",
        pages=[
            PageMetadata(
                page_number=1,
                page_count=1,
                width=8.5,
                height=11,
                unit="inch",
                source_text="功能货币为美元，列报货币为人民币。",
            )
        ],
    )

    contexts = FinancialExtractionService._currency_contexts(document)

    # A page naming two distinct currencies must not silently collapse to a single
    # guessed currency - this is a correctness improvement over the prior behavior
    # (only 人民币 was recognized, so this page previously guessed CNY alone).
    assert 1 not in contexts


def test_single_currency_name_still_resolves_page_context() -> None:
    document = CanonicalDocument(
        document_id="doc-single-currency",
        filename="annual-report.pdf",
        status="normalizing",
        pages=[
            PageMetadata(
                page_number=1,
                page_count=1,
                width=8.5,
                height=11,
                unit="inch",
                source_text="列報貨幣為港元。",
            )
        ],
    )

    contexts = FinancialExtractionService._currency_contexts(document)

    assert contexts[1] == "HKD"


def test_date_context_re_recognizes_zhi_and_niandu() -> None:
    assert DATE_CONTEXT_RE.search("截至2024年12月31日止年度")
    assert DATE_CONTEXT_RE.search("年度报告")


def test_bare_zhi_character_not_added_to_date_context() -> None:
    # Confirmed decision: bare 自/至/止 are excluded because they cause a real
    # false positive - an "automation ID" cell must still classify as an identifier,
    # not get suppressed by DATE_CONTEXT_RE matching a bare "自" inside "自动化".
    label_context = "自动化编号"
    assert not DATE_CONTEXT_RE.search(label_context)
    assert IDENTIFIER_CONTEXT_RE.search(label_context)
