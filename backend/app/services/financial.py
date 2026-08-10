"""Financial page selection, normalization, and deterministic validation."""

from __future__ import annotations

import csv
import hashlib
import json
import re
from decimal import Decimal, InvalidOperation
from io import BytesIO, StringIO
from typing import Literal

from app.core.enums import FinancialPageDisposition
from app.schemas.financial import (
    FinancialCell,
    FinancialClassificationEvidence,
    FinancialClassificationResult,
    FinancialContentItem,
    FinancialPageClassification,
    FinancialResult,
    FinancialTable,
    FinancialValidationIssue,
    FinancialValidationReport,
    NormalizedFinancialValue,
)
from app.schemas.page import CanonicalDocument, TableCell, TextBlock

CURRENCY_SYMBOLS = {"$": "USD", "€": "EUR", "£": "GBP", "¥": "JPY"}
CURRENCY_SYMBOL_CANDIDATES = {
    "$": ("USD",),
    "\u20ac": ("EUR",),
    "\u00a3": ("GBP",),
    "\u00a5": ("CNY", "JPY"),
    "\u20b9": ("INR",),
}
CURRENCY_CODE_RE = re.compile(
    r"\b(USD|EUR|GBP|JPY|CNY|RMB|INR|AED|SAR|CAD|AUD|CHF)\b",
    re.IGNORECASE,
)
CURRENCY_NAME_CONTEXT = {
    "\u4eba\u6c11\u5e01": "CNY",
    "\u4eba\u6c11\u5e63": "CNY",
    "renminbi": "CNY",
    "yuan": "CNY",
    "\u5186": "JPY",
    "yen": "JPY",
    "\u6e2f\u5143": "HKD",
    "\u6e2f\u5e63": "HKD",
    "\u6e2f\u5e01": "HKD",
    "\u65b0\u53f0\u5e63": "TWD",
    "\u65b0\u53f0\u5e01": "TWD",
    "\u6fb3\u9580\u5143": "MOP",
    "\u7f8e\u5143": "USD",
    "\u6b50\u5143": "EUR",
    "\u6b27\u5143": "EUR",
    "\u82f1\u938a": "GBP",
    "\u82f1\u9551": "GBP",
    "\u65e5\u5143": "JPY",
    "\u65e5\u5713": "JPY",
}

FINANCIAL_TERMS = {
    "asset",
    "balance",
    "cash",
    "credit",
    "debit",
    "expense",
    "income",
    "invoice",
    "liability",
    "payable",
    "profit",
    "revenue",
    "subtotal",
    "tax",
    "total",
}
FINANCIAL_CONTEXT_TERMS = FINANCIAL_TERMS | {
    "account",
    "amount",
    "capital",
    "closing",
    "cost",
    "currency",
    "equity",
    "gross",
    "interest",
    "net",
    "opening",
    "payment",
    "period",
    "statement",
    "turnover",
    "year ended",
}
# Shared CJK/Arabic monetary vocabulary - single source of truth for MONETARY_CONTEXT_RE
# (per-cell semantic typing) below, reused here for page-level classification so the two
# never drift out of sync again.
MONETARY_TERMS_LATIN = (
    "amount", "asset", "liability", "balance", "debit", "credit", "income", "expense",
    "revenue", "cost", "price", "subtotal", "total", "tax", "payment", "refund",
    "opening", "closing", "payable", "receivable",
)
MONETARY_TERMS_CJK = (
    "金额", "金額", "余额", "餘額", "支出", "收入", "单价", "單價", "合计", "合計",
    "税额", "稅額", "借方", "贷方", "貸方", "应收", "应付", "應收", "應付",
    "费用", "費用", "成本", "利息", "溢利", "亏损", "虧損", "贷款", "貸款", "借款",
    "股息", "净额", "淨額",
)
MONETARY_TERMS_ARABIC = ("مبلغ", "رصيد", "إجمالي", "ضريبة", "مدين", "دائن")
# Page-level financial vocabulary (broader than the per-cell monetary terms above - also
# needs document-identifying terms like "balance sheet"/"annual report"). Covers both
# Simplified and Traditional Chinese since real-world filers use either.
FINANCIAL_TERMS_CJK = MONETARY_TERMS_CJK + (
    "资产", "資產", "负债", "負債", "现金", "現金", "利润", "利潤", "损益", "損益",
    "资产负债表", "資產負債表", "财务报表", "財務報表", "年度报告", "年度報告",
    "股东权益", "股東權益", "营业收入", "營業收入", "毛利", "净利", "淨利",
    # Balance sheet
    "流动资产", "流動資產", "存货", "存貨", "商誉", "商譽", "股本", "权益", "權益",
    # Income statement
    "折旧", "折舊", "摊销", "攤銷", "每股收益", "每股盈利",
    # Cash flow statement
    "经营活动", "經營活動", "投资活动", "投資活動", "筹资活动", "籌資活動",
    # Notes / governance
    "董事", "联营公司", "聯營公司", "合营公司", "合營公司", "关联方", "關聯方",
    "审计报告", "審計報告", "汇率", "匯率",
    # User-confirmed set from the earlier document review
    "所得税", "所得稅", "减值", "減值", "盈利", "股东", "股東", "每股",
    "持续经营业务", "持續經營業務", "归属", "歸屬", "综合", "綜合", "合并", "合併",
)
FINANCIAL_CONTEXT_TERMS_CJK = FINANCIAL_TERMS_CJK
# CJK has no whitespace between words, so \b (a \w/non-\w transition) never falls between
# two adjacent CJK characters - a \b-wrapped search for "資產" inside "資產負債表" never
# matches. Plain substring alternation (no \b) is required for these terms; this mirrors
# the strategy MONETARY_CONTEXT_RE already uses below, which is why it "just works" for
# CJK while the \b-wrapped FINANCIAL_TERMS loop does not.
FINANCIAL_TERMS_CJK_RE = re.compile("|".join(re.escape(term) for term in FINANCIAL_TERMS_CJK))
FINANCIAL_CONTEXT_TERMS_CJK_RE = re.compile(
    "|".join(re.escape(term) for term in FINANCIAL_CONTEXT_TERMS_CJK)
)
NUMBER_RE = re.compile(r"^[+-]?\d[\d., ]*$")
# Full-width digits already parse correctly via Python's Unicode-aware \d/Decimal() -
# only the CJK-typeset punctuation needs an explicit mapping to ASCII before the rest
# of the numeric-parsing logic (which is ASCII-only) can see it.
FULLWIDTH_PUNCTUATION_TRANSLATION = str.maketrans({"，": ",", "．": "."})
# CJK numeral magnitude suffixes: 億/亿 = hundred million (10^8), 萬/万 = ten thousand
# (10^4). Deliberately restricted to a single plain integer/decimal followed by exactly
# one suffix (and an optional trailing currency unit) - comma-grouped values and
# compound forms like "2億3000萬" (implied addition) are out of scope and fall through
# unchanged to the existing ambiguous/unparsed handling.
CJK_MAGNITUDE_RE = re.compile(
    r"^(?P<number>\d+(?:\.\d+)?)\s*(?P<magnitude>億|亿|萬|万)\s*(?:元|圆|圓)?$"
)
CJK_MAGNITUDE_MULTIPLIERS = {"億": 10**8, "亿": 10**8, "萬": 10**4, "万": 10**4}
# Minimum uniformly-comma-grouped cells required before _infer_decimal_separator's
# weak-evidence fallback treats the pattern as "this table uses comma grouping" rather
# than a coincidence in a small/sparse table.
WEAK_EVIDENCE_MIN_CELLS = 2
PERCENTAGE_RANGE_RE = re.compile(
    r"^\s*[+-]?\d+(?:[.,]\d+)?\s*%\s*(?:-|\u2013|\u2014|~|to|\u81f3|\u5230)\s*"
    r"[+-]?\d+(?:[.,]\d+)?\s*%\s*$",
    re.IGNORECASE,
)
DATE_OR_TIME_RE = re.compile(
    r"(?:\b\d{4}[-/.]\d{1,2}[-/.]\d{1,2}\b|"
    r"\b\d{1,2}[-/.]\d{1,2}[-/.]\d{2,4}\b|"
    r"\d{4}\s*\u5e74\s*\d{1,2}\s*\u6708\s*\d{1,2}\s*\u65e5|"
    r"\b\d{1,2}:\d{2}(?::\d{2})?\b)"
)
DATE_CONTEXT_RE = re.compile(
    r"(?:date|time|period|from|through|\u65e5\u671f|\u65f6\u95f4|\u671f\u95f4|\u671f\u9593|\u622a\u81f3|\u5e74\u5ea6|"
    r"\u062a\u0627\u0631\u064a\u062e|\u0648\u0642\u062a|\u0641\u062a\u0631\u0629)",
    re.IGNORECASE,
)
PHONE_CONTEXT_RE = re.compile(
    r"(?:phone|telephone|mobile|contact|hotline|tel\.?|\u7535\u8bdd|\u624b\u673a|"
    r"\u8054\u7cfb|\u70ed\u7ebf|\u96fb\u8a71|\u624b\u6a5f|\u0627\u0644\u0647\u0627\u062a\u0641|\u062c\u0648\u0627\u0644)",
    re.IGNORECASE,
)
ACCOUNT_CONTEXT_RE = re.compile(
    r"(?:account\s*(?:number|no\.?|#)|iban|card\s*(?:number|no\.?)|"
    r"\u8d26\u6237\u53f7\u7801|\u8d26\u53f7|\u5e10\u6237\u53f7\u7801|\u5e10\u53f7|"
    r"\u5361\u53f7|\u8cec\u6236\u865f\u78bc|\u5e33\u865f|\u0631\u0642\u0645\s*\u0627\u0644\u062d\u0633\u0627\u0628)",
    re.IGNORECASE,
)
IDENTIFIER_CONTEXT_RE = re.compile(
    r"(?:transaction|reference|invoice\s*(?:number|no\.?)|customer\s*(?:number|id)|"
    r"identifier|\bid\b|\u4ea4\u6613\u7f16\u53f7|\u53d1\u7968\u53f7|\u5ba2\u6237\u7f16\u53f7|"
    r"\u8bc1\u4ef6\u53f7\u7801|\u7f16\u53f7|\u7de8\u865f|\u0631\u0642\u0645\s*\u0627\u0644\u0645\u0639\u0627\u0645\u0644\u0629)",
    re.IGNORECASE,
)
MEASUREMENT_UNIT_RE = re.compile(
    r"(?:mg\s*/\s*l|g\s*/\s*l|mmol\s*/\s*l|mol\s*/\s*l|"
    r"mmhg|kpa|bpm|kg|cm|mm|\u00b0\s*c|\u00b0\s*f|"
    r"(?:[x\u00d7]\s*10(?:\s*\^?\s*\d+)?)?\s*/\s*l|"
    r"\u6b21\s*/\s*\u5206\u949f|\u5c81|\u6b72)",
    re.IGNORECASE,
)
QUANTITY_CONTEXT_RE = re.compile(
    r"(?:quantity|qty\.?|units?|count|\u6570\u91cf|\u6578\u91cf|\u4ef6\u6570|\u4ef6\u6578|\u0639\u062f\u062f|\u0627\u0644\u0643\u0645\u064a\u0629)",
    re.IGNORECASE,
)
MONETARY_CONTEXT_RE = re.compile(
    "|".join(
        re.escape(term)
        for term in MONETARY_TERMS_LATIN + MONETARY_TERMS_CJK + MONETARY_TERMS_ARABIC
    ),
    re.IGNORECASE,
)
KEY_VALUE_RE = re.compile(r"^\s*([^:\n：]{1,120})\s*[:：]\s*(.+?)\s*$", re.DOTALL)
LIST_ITEM_RE = re.compile(r"^\s*(?:[-•‣▪◦]|\d+[.)])\s+(.+)$", re.DOTALL)
FINANCIAL_VALUE_RE = re.compile(
    r"(?:[$€£¥₹]|\b(?:USD|EUR|GBP|JPY|CNY|RMB|INR|AED|SAR|CAD|AUD|CHF)\b)"
    r"\s*[+-]?\d|[+-]?\d[\d., ]*\s*%",
    re.IGNORECASE,
)
SPREADSHEET_FORMULA_PREFIXES = ("=", "+", "-", "@", "＝", "＋", "－", "＠")


def _spreadsheet_safe(value: str) -> str:
    stripped = value.lstrip()
    return "'" + value if stripped.startswith(SPREADSHEET_FORMULA_PREFIXES) else value


def financial_result_sha256(result: FinancialResult) -> str:
    """Return the canonical digest used to bind review decisions to one result."""
    payload = json.dumps(
        result.model_dump(mode="json"),
        ensure_ascii=False,
        sort_keys=True,
        separators=(",", ":"),
    ).encode("utf-8")
    return hashlib.sha256(payload).hexdigest()


def financial_result_csv(result: FinancialResult) -> str:
    """Create a flat, formula-injection-safe CSV with full cell provenance."""
    output = StringIO(newline="")
    writer = csv.writer(output)
    writer.writerow(
        [
            "document_id",
            "table_id",
            "source_page",
            "row_index",
            "column_index",
            "raw_text",
            "normalized_value",
            "value_type",
            "semantic_type",
            "currency",
            "currency_candidates",
            "currency_source",
            "currency_ambiguous",
            "requires_normalized_correction",
            "requires_currency_correction",
            "review_reason_code",
            "cell_origin",
            "source_block_ids",
            "reconciliation_candidate_id",
            "table_integrity_status",
            "review_required",
        ]
    )
    for table in result.tables:
        for cell in table.cells:
            writer.writerow(
                [
                    _spreadsheet_safe(result.document_id),
                    _spreadsheet_safe(table.table_id),
                    cell.source_page,
                    cell.row_index,
                    cell.column_index,
                    _spreadsheet_safe(cell.value.raw_text),
                    _spreadsheet_safe(str(cell.value.normalized_value))
                    if cell.value.normalized_value is not None
                    else "",
                    _spreadsheet_safe(cell.value.value_type),
                    _spreadsheet_safe(cell.value.semantic_type),
                    _spreadsheet_safe(cell.value.currency or ""),
                    _spreadsheet_safe(",".join(cell.value.currency_candidates)),
                    _spreadsheet_safe(cell.value.currency_source or ""),
                    cell.value.currency_ambiguous,
                    cell.value.requires_normalized_correction,
                    cell.value.requires_currency_correction,
                    _spreadsheet_safe(cell.value.review_reason_code or ""),
                    _spreadsheet_safe(cell.origin),
                    _spreadsheet_safe(",".join(cell.source_block_ids)),
                    _spreadsheet_safe(cell.reconciliation_candidate_id or ""),
                    _spreadsheet_safe(table.integrity_status),
                    cell.review_required
                    or cell.value.requires_normalized_correction
                    or cell.value.requires_currency_correction,
                ]
            )
    return output.getvalue()


def financial_result_xlsx(result: FinancialResult) -> bytes:
    """Create an XLSX workbook with one sheet per financial table and audit sheets."""
    from openpyxl import Workbook
    from openpyxl.comments import Comment

    workbook = Workbook()
    manifest = workbook.active
    if manifest is None:
        raise RuntimeError("The workbook did not create its manifest sheet.")
    manifest.title = "Manifest"
    manifest.append(["Document ID", _spreadsheet_safe(result.document_id)])
    manifest.append(["Schema version", _spreadsheet_safe(result.schema_version)])
    manifest.append(["Processing version", _spreadsheet_safe(result.processing_version)])
    manifest.append(["Source pages", result.source_page_count])
    manifest.append(
        ["Selected pages", _spreadsheet_safe(", ".join(map(str, result.selected_pages)))]
    )

    used_names = {manifest.title}
    for index, table in enumerate(result.tables, start=1):
        base = re.sub(r"[\\/*?:\[\]]", "_", f"Table {index} {table.financial_type}")[:31]
        title = base or f"Table {index}"
        suffix = 1
        while title in used_names:
            suffix += 1
            title = f"{base[:27]} {suffix}"[:31]
        used_names.add(title)
        sheet = workbook.create_sheet(title)
        for cell in table.cells:
            target = sheet.cell(row=cell.row_index + 1, column=cell.column_index + 1)
            normalized = cell.value.normalized_value
            target.value = (
                _spreadsheet_safe(str(normalized))
                if normalized is not None
                else _spreadsheet_safe(cell.value.raw_text)
            )
            provenance = [
                f"Source: {cell.value.raw_text}",
                f"Page: {cell.source_page}",
                f"Origin: {cell.origin}",
            ]
            if cell.source_block_ids:
                provenance.append(f"Source blocks: {', '.join(cell.source_block_ids)}")
            if cell.reconciliation_candidate_id:
                provenance.append(
                    f"Reconciliation: {cell.reconciliation_candidate_id}"
                )
            if normalized is not None or cell.origin == "reconstructed":
                target.comment = Comment(
                    "\n".join(provenance),
                    "CareTranslate Studio",
                )

    validation = workbook.create_sheet("Validation")
    validation.append(["Severity", "Code", "Page", "Table", "Cell", "Message"])
    for issue in result.validation.issues:
        validation.append(
            [
                _spreadsheet_safe(issue.severity),
                _spreadsheet_safe(issue.code),
                issue.page_number,
                _spreadsheet_safe(issue.table_id or ""),
                _spreadsheet_safe(issue.cell_id or ""),
                _spreadsheet_safe(issue.message),
            ]
        )
    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def selected_page_ranges(page_numbers: list[int], maximum_span: int) -> list[tuple[int, int]]:
    """Group selected pages into contiguous ranges capped at the DI range limit."""
    if maximum_span < 1:
        raise ValueError("maximum_span must be positive")
    pages = sorted(set(page_numbers))
    if not pages:
        return []
    ranges: list[tuple[int, int]] = []
    start = previous = pages[0]
    for page in pages[1:]:
        would_exceed = page - start + 1 > maximum_span
        if page != previous + 1 or would_exceed:
            ranges.append((start, previous))
            start = page
        previous = page
    ranges.append((start, previous))
    return ranges


class FinancialCandidateSelector:
    """Convert provider classifications into conservative, recall-first page decisions."""

    def __init__(
        self,
        *,
        financial_labels: set[str],
        include_confidence: float,
        exclude_confidence: float,
        adjacent_pages: int,
    ) -> None:
        self.financial_labels = {label.casefold() for label in financial_labels}
        self.include_confidence = include_confidence
        self.exclude_confidence = exclude_confidence
        self.adjacent_pages = adjacent_pages
        policy_payload = json.dumps(
            {
                "algorithm": "recall-first-2",
                "financial_labels": sorted(self.financial_labels),
                "include_confidence": include_confidence,
                "exclude_confidence": exclude_confidence,
                "adjacent_pages": adjacent_pages,
            },
            sort_keys=True,
            separators=(",", ":"),
        )
        self.policy_fingerprint = hashlib.sha256(policy_payload.encode()).hexdigest()

    def from_provider(
        self,
        *,
        document_id: str,
        classifier_id: str,
        classifier_version: str,
        source_page_count: int,
        raw: dict[str, object],
    ) -> FinancialClassificationResult:
        candidates_by_page: dict[int, list[tuple[str, float]]] = {}
        documents = raw.get("documents")
        if isinstance(documents, list):
            for item in documents:
                if not isinstance(item, dict):
                    continue
                label = str(item.get("doc_type") or item.get("docType") or "unknown")
                confidence = float(item.get("confidence") or 0)
                regions = item.get("bounding_regions") or item.get("boundingRegions") or []
                if not isinstance(regions, list):
                    continue
                for region in regions:
                    if not isinstance(region, dict):
                        continue
                    page_number = int(
                        region.get("page_number") or region.get("pageNumber") or 0
                    )
                    if 1 <= page_number <= source_page_count:
                        candidates_by_page.setdefault(page_number, []).append(
                            (label, confidence)
                        )

        pages: list[FinancialPageClassification] = []
        for page_number in range(1, source_page_count + 1):
            candidates = sorted(
                candidates_by_page.get(page_number, []),
                key=lambda item: item[1],
                reverse=True,
            )
            financial_candidates = [
                item for item in candidates if item[0].casefold() in self.financial_labels
            ]
            non_financial_candidates = [
                item for item in candidates if item[0].casefold() not in self.financial_labels
            ]
            included_financial = [
                item for item in financial_candidates if item[1] >= self.include_confidence
            ]
            excluded_non_financial = [
                item
                for item in non_financial_candidates
                if item[1] >= self.exclude_confidence
            ]
            if included_financial:
                label, confidence = included_financial[0]
                disposition = FinancialPageDisposition.FINANCIAL
                selected = True
                review_required = bool(excluded_non_financial)
                reasons = ["financial_classifier_match"]
                if review_required:
                    reasons.append("conflicting_non_financial_evidence")
            elif financial_candidates:
                label, confidence = financial_candidates[0]
                disposition = FinancialPageDisposition.UNCERTAIN
                selected = True
                review_required = True
                reasons = ["financial_evidence_below_include_threshold"]
                if excluded_non_financial:
                    reasons.append("conflicting_non_financial_evidence")
            elif excluded_non_financial:
                label, confidence = excluded_non_financial[0]
                disposition = FinancialPageDisposition.NON_FINANCIAL
                selected = False
                review_required = False
                reasons = ["high_confidence_non_financial"]
            else:
                label, confidence = candidates[0] if candidates else ("unknown", 0.0)
                disposition = FinancialPageDisposition.UNCERTAIN
                selected = True
                review_required = True
                reasons = ["classification_uncertain"]
            pages.append(
                FinancialPageClassification(
                    page_number=page_number,
                    label=label,
                    confidence=confidence,
                    disposition=disposition,
                    selected=selected,
                    review_required=review_required,
                    reasons=reasons,
                    candidates=[
                        FinancialClassificationEvidence(
                            label=candidate_label,
                            confidence=candidate_confidence,
                        )
                        for candidate_label, candidate_confidence in candidates
                    ],
                    source="azure_custom_classifier",
                )
            )

        self._include_adjacent_pages(pages)
        return FinancialClassificationResult(
            document_id=document_id,
            classifier_id=classifier_id,
            classifier_version=classifier_version,
            selection_policy_fingerprint=self.policy_fingerprint,
            source_page_count=source_page_count,
            pages=pages,
        )

    def from_layout(
        self,
        *,
        document: CanonicalDocument,
        source_page_count: int,
    ) -> FinancialClassificationResult:
        tables_by_page = {
            region.page_number
            for table in document.tables
            for cell in table.cells
            for region in cell.bounding_regions
        }
        pages: list[FinancialPageClassification] = []
        metadata = {page.page_number: page for page in document.pages}
        for page_number in range(1, source_page_count + 1):
            text = (
                metadata[page_number].source_text.casefold()
                if page_number in metadata
                else ""
            )
            term_hits = sum(
                1
                for term in FINANCIAL_TERMS
                if re.search(rf"\b{re.escape(term)}\b", text)
            ) + len(FINANCIAL_TERMS_CJK_RE.findall(text))
            has_table = page_number in tables_by_page
            structural_value_hits = len(FINANCIAL_VALUE_RE.findall(text))
            explicit_currency_hits = len(CURRENCY_CODE_RE.findall(text)) + sum(
                text.count(symbol) for symbol in CURRENCY_SYMBOL_CANDIDATES
            )
            measurement_hits = len(MEASUREMENT_UNIT_RE.findall(text))
            selected = has_table and (
                term_hits > 0
                or explicit_currency_hits > 0
            )
            uncertain = has_table and not selected
            disposition = (
                FinancialPageDisposition.FINANCIAL
                if selected
                else FinancialPageDisposition.UNCERTAIN
                if uncertain
                else FinancialPageDisposition.NON_FINANCIAL
            )
            pages.append(
                FinancialPageClassification(
                    page_number=page_number,
                    label=(
                        "financial_table"
                        if selected
                        else "table_unknown"
                        if uncertain
                        else "non_financial"
                    ),
                    confidence=0.75 if selected else 0.5 if uncertain else 0.8,
                    disposition=disposition,
                    selected=selected or uncertain,
                    review_required=uncertain,
                    reasons=[
                        "layout_table_and_financial_terms"
                        if term_hits
                        else "layout_table_and_currency_values"
                    ]
                    if selected
                    else [
                        "layout_measurement_table_requires_review"
                        if measurement_hits and structural_value_hits
                        else "layout_table_requires_review"
                    ]
                    if uncertain
                    else ["no_financial_layout_signal"],
                    source="layout_rules",
                )
            )
        self._include_adjacent_pages(pages)
        return FinancialClassificationResult(
            document_id=document.document_id,
            classifier_id="layout-rules",
            classifier_version="2.3",
            selection_policy_fingerprint=self.policy_fingerprint,
            source_page_count=source_page_count,
            pages=pages,
        )

    def _include_adjacent_pages(self, pages: list[FinancialPageClassification]) -> None:
        selected = {page.page_number for page in pages if page.selected}
        adjacent: set[int] = set()
        for page_number in selected:
            for offset in range(1, self.adjacent_pages + 1):
                adjacent.update({page_number - offset, page_number + offset})
        for page in pages:
            if page.page_number in adjacent and not page.selected:
                page.selected = True
                page.review_required = True
                page.disposition = FinancialPageDisposition.UNCERTAIN
                page.reasons.append("adjacent_to_financial_page")
                page.source = "adjacent_page"


class FinancialExtractionService:
    """Build normalized financial artifacts from the immutable canonical extraction."""

    @staticmethod
    def _content_type(
        block: TextBlock,
    ) -> tuple[
        Literal["heading", "paragraph", "key_value", "list_item"],
        re.Match[str] | None,
    ]:
        role = (block.role or "").casefold()
        if role in {"title", "sectionheading"}:
            return "heading", None
        if LIST_ITEM_RE.match(block.source_text):
            return "list_item", None
        key_value = KEY_VALUE_RE.match(block.source_text)
        if key_value:
            return "key_value", key_value
        return "paragraph", None

    @staticmethod
    def _has_financial_signal(text: str) -> bool:
        folded = text.casefold()
        if any(
            re.search(rf"\b{re.escape(term)}\b", folded)
            for term in FINANCIAL_CONTEXT_TERMS
        ):
            return True
        if FINANCIAL_CONTEXT_TERMS_CJK_RE.search(text):
            return True
        return bool(FINANCIAL_VALUE_RE.search(text))

    @staticmethod
    def _span_position(block: TextBlock) -> int:
        return min((span.offset for span in block.spans), default=block.reading_order * 10_000)

    def _content_items(
        self,
        document: CanonicalDocument,
        classification: FinancialClassificationResult,
        tables: list[FinancialTable],
    ) -> list[FinancialContentItem]:
        """Project selected financial structure without flattening source formats."""

        selected_pages = set(classification.selected_pages)
        if not selected_pages:
            return []
        decisions = {page.page_number: page for page in classification.pages}
        drafts: list[tuple[int, int, int, FinancialContentItem]] = []

        financial_block_ids: set[str] = set()
        anchors_by_page: dict[int, list[int]] = {}
        headings_by_page: dict[int, list[tuple[int, str]]] = {}

        for block in document.blocks:
            source_pages = sorted(
                {
                    region.page_number
                    for region in block.bounding_regions
                    if region.page_number in selected_pages
                }
            )
            if not source_pages:
                continue
            item_type, _ = self._content_type(block)
            position = self._span_position(block)
            if item_type == "heading":
                for page in source_pages:
                    headings_by_page.setdefault(page, []).append(
                        (position, block.block_id)
                    )
            if self._has_financial_signal(block.source_text):
                financial_block_ids.add(block.block_id)
                for page in source_pages:
                    anchors_by_page.setdefault(page, []).append(position)

        for source_table in document.tables:
            for page in selected_pages:
                page_spans = [
                    span.offset
                    for cell in source_table.cells
                    if any(
                        region.page_number == page
                        for region in cell.bounding_regions
                    )
                    for span in cell.spans
                ]
                if page_spans or any(
                    region.page_number == page
                    for region in source_table.bounding_regions
                ):
                    anchors_by_page.setdefault(page, []).append(
                        min(page_spans, default=10**9)
                    )

        supporting_heading_ids: set[str] = set()
        for page, anchors in anchors_by_page.items():
            candidates = sorted(headings_by_page.get(page, []))
            for anchor in anchors:
                preceding = [
                    block_id for position, block_id in candidates if position <= anchor
                ]
                if preceding:
                    supporting_heading_ids.add(preceding[-1])
        included_block_ids = financial_block_ids | supporting_heading_ids


        for block in document.blocks:
            source_pages = sorted(
                {
                    region.page_number
                    for region in block.bounding_regions
                    if region.page_number in selected_pages
                }
            )
            if not source_pages:
                continue
            page_decisions = [decisions[page] for page in source_pages]
            item_type, key_value = self._content_type(block)
            financial_signal = self._has_financial_signal(block.source_text)
            if block.block_id not in included_block_ids:
                continue

            uncertain = all(
                page.disposition == FinancialPageDisposition.UNCERTAIN
                for page in page_decisions
            )
            block_relevance: Literal["financial", "supporting", "uncertain"]
            if uncertain:
                block_relevance = "uncertain"
            elif financial_signal:
                block_relevance = "financial"
            else:
                block_relevance = "supporting"
            block_warnings = list(block.warnings)
            translated_key_value = (
                KEY_VALUE_RE.match(block.translated_text)
                if item_type == "key_value" and block.translated_text
                else None
            )
            translation_format_review = (
                item_type == "key_value"
                and bool(block.translated_text)
                and translated_key_value is None
            )
            if block_relevance == "uncertain":
                block_warnings.append(
                    "Financial relevance requires reviewer confirmation."
                )
            if translation_format_review:
                block_warnings.append(
                    "Translated key-value formatting requires reviewer confirmation."
                )
            block_item = FinancialContentItem(
                item_id=block.block_id,
                item_type=item_type,
                reading_order=1,
                source_pages=source_pages,
                source_language=block.source_language,
                source_text=block.source_text,
                translated_text=block.translated_text,
                role=block.role,
                key=key_value.group(1).strip() if key_value else None,
                value=key_value.group(2).strip() if key_value else None,
                translated_key=(
                    translated_key_value.group(1).strip()
                    if translated_key_value
                    else None
                ),
                translated_value=(
                    translated_key_value.group(2).strip()
                    if translated_key_value
                    else None
                ),
                relevance=block_relevance,
                bounding_regions=[
                    region
                    for region in block.bounding_regions
                    if region.page_number in selected_pages
                ],
                review_required=block.review_required
                or translation_format_review
                or any(page.review_required for page in page_decisions),
                warnings=block_warnings,
            )
            drafts.append(
                (
                    source_pages[0],
                    self._span_position(block),
                    0,
                    block_item,
                )
            )

        raw_tables = {
            source_table.table_id: source_table for source_table in document.tables
        }
        for financial_table in tables:
            raw_table = raw_tables.get(financial_table.table_id)
            table_decisions = [
                decisions[page]
                for page in financial_table.source_pages
                if page in decisions and page in selected_pages
            ]
            definitive = any(
                page.disposition == FinancialPageDisposition.FINANCIAL
                for page in table_decisions
            )
            table_relevance: Literal["financial", "supporting", "uncertain"] = (
                "financial" if definitive else "uncertain"
            )
            languages = (
                {
                    cell.source_language
                    for cell in raw_table.cells
                    if cell.source_language not in {"", "und"}
                }
                if raw_table is not None
                else set()
            )
            if not languages:
                source_language = "und"
            elif len(languages) == 1:
                source_language = next(iter(languages))
            else:
                source_language = "mixed"
            table_warnings: list[str] = []
            if not financial_table.complete:
                table_warnings.append(
                    "The provider table is incomplete across its source pages."
                )
            if financial_table.classification_override_pages:
                table_warnings.append(
                    "The table crosses a financial page-selection boundary."
                )
            if table_relevance == "uncertain":
                table_warnings.append(
                    "Financial relevance requires reviewer confirmation."
                )
            if financial_table.integrity_status == "reconciled":
                table_warnings.append(
                    "The effective table includes aligned source blocks that require "
                    "reviewer acceptance."
                )
            span_position = (
                min(
                    (
                        span.offset
                        for cell in raw_table.cells
                        for span in cell.spans
                    ),
                    default=10**9,
                )
                if raw_table is not None
                else 10**9
            )
            table_item = FinancialContentItem(
                item_id=f"table:{financial_table.table_id}",
                item_type="table",
                reading_order=1,
                source_pages=financial_table.source_pages,
                source_language=source_language,
                table_id=financial_table.table_id,
                relevance=table_relevance,
                bounding_regions=list(raw_table.bounding_regions)
                if raw_table is not None
                else [],
                review_required=financial_table.review_required
                or any(page.review_required for page in table_decisions),
                warnings=table_warnings,
            )
            drafts.append(
                (
                    financial_table.source_pages[0],
                    span_position,
                    1,
                    table_item,
                )
            )

        drafts.sort(key=lambda draft: (draft[0], draft[1], draft[2], draft[3].item_id))
        items = [draft[3] for draft in drafts]
        for reading_order, item in enumerate(items, start=1):
            item.reading_order = reading_order
        return items

    def build(
        self,
        document: CanonicalDocument,
        classification: FinancialClassificationResult,
        *,
        processing_version: str,
    ) -> FinancialResult:
        selected = set(classification.selected_pages)
        page_labels = {page.page_number: page.label for page in classification.pages}
        page_decisions = {
            page.page_number: page for page in classification.pages
        }
        currency_contexts = self._currency_contexts(document)
        tables: list[FinancialTable] = []
        for table in document.tables:
            cell_source_pages = {
                region.page_number
                for cell in table.cells
                for region in cell.bounding_regions
            }
            provider_source_pages = cell_source_pages | {
                region.page_number for region in table.bounding_regions
            }
            if not provider_source_pages:
                provider_source_pages = {1}
            if not (provider_source_pages & selected):
                continue
            decimal_separator = self._infer_decimal_separator(table.cells)
            cells = [
                self._cell(
                    cell,
                    table_cells=table.cells,
                    decimal_separator=decimal_separator,
                    currency_context=currency_contexts.get(self._cell_page(cell)),
                )
                for cell in table.cells
            ]
            # A provider table is atomic. Once any of its pages is selected, keep every
            # returned cell instead of retaining its grid dimensions while dropping
            # cells on a competing/non-selected page.
            missing_source_pages = sorted(provider_source_pages - cell_source_pages)
            classification_override_pages = sorted(cell_source_pages - selected)
            source_pages = sorted(provider_source_pages)
            currencies = sorted(
                {cell.value.currency for cell in cells if cell.value.currency is not None}
            )
            definitive_financial = any(
                page_decisions[page].disposition == FinancialPageDisposition.FINANCIAL
                for page in source_pages
                if page in page_decisions
            )
            reconciled = table.integrity_status == "reconciled"
            tables.append(
                FinancialTable(
                    table_id=table.table_id,
                    source_pages=source_pages,
                    financial_type=page_labels.get(
                        min(provider_source_pages & selected), "financial_table"
                    ),
                    row_count=table.row_count,
                    column_count=table.column_count,
                    currencies=currencies,
                    cells=cells,
                    complete=not missing_source_pages,
                    missing_source_pages=missing_source_pages,
                    classification_override_pages=classification_override_pages,
                    review_required=bool(missing_source_pages)
                    or bool(classification_override_pages)
                    or reconciled
                    or any(cell.review_required for cell in cells),
                    merge_policy=(
                        "provider_table_with_reconciliation_v1"
                        if reconciled
                        else "provider_table_atomic_v1"
                    ),
                    provider_row_count=table.provider_row_count,
                    provider_column_count=table.provider_column_count,
                    integrity_status=table.integrity_status,
                    reconciliation_candidate_id=table.reconciliation_candidate_id,
                    relevance="financial" if definitive_financial else "uncertain",
                )
            )

        content_items = self._content_items(document, classification, tables)
        issues = self._validate(classification, tables, content_items)
        report = FinancialValidationReport(
            document_id=document.document_id,
            issue_count=len(issues),
            error_count=sum(issue.severity == "error" for issue in issues),
            warning_count=sum(issue.severity == "warning" for issue in issues),
            review_required=bool(issues)
            or any(page.review_required for page in classification.pages if page.selected),
            issues=issues,
        )
        return FinancialResult(
            document_id=document.document_id,
            processing_version=processing_version,
            source_page_count=classification.source_page_count,
            selected_pages=classification.selected_pages,
            uncertain_pages=[
                page.page_number
                for page in classification.pages
                if page.disposition == FinancialPageDisposition.UNCERTAIN
            ],
            content_items=content_items,
            tables=tables,
            validation=report,
            reconciliation_candidate_ids=sorted(
                {
                    table.reconciliation_candidate_id
                    for table in tables
                    if table.reconciliation_candidate_id is not None
                }
            ),
        )

    @staticmethod
    def _cell_page(cell: TableCell) -> int:
        return cell.bounding_regions[0].page_number if cell.bounding_regions else 1

    def _cell(
        self,
        cell: TableCell,
        *,
        table_cells: list[TableCell],
        decimal_separator: str | None,
        currency_context: str | None,
    ) -> FinancialCell:
        semantic_type = self._semantic_type(cell, table_cells)
        value = self._normalize(
            cell.content,
            semantic_type=semantic_type,
            decimal_separator=decimal_separator,
            currency_context=currency_context,
        )
        warnings = list(cell.warnings)
        if value.requires_currency_correction:
            warnings.append("Currency symbol is ambiguous and requires reviewer confirmation.")
        return FinancialCell(
            cell_id=cell.cell_id,
            row_index=cell.row_index,
            column_index=cell.column_index,
            row_span=cell.row_span,
            column_span=cell.column_span,
            kind=cell.kind,
            source_language=cell.source_language,
            translated_text=cell.translated_content,
            value=value,
            source_page=self._cell_page(cell),
            bounding_regions=cell.bounding_regions,
            review_required=cell.review_required
            or value.requires_normalized_correction
            or value.requires_currency_correction,
            warnings=warnings,
            origin=cell.origin,
            source_block_ids=cell.source_block_ids,
            reconciliation_candidate_id=cell.reconciliation_candidate_id,
        )

    @classmethod
    def _semantic_type(
        cls,
        cell: TableCell,
        table_cells: list[TableCell],
    ) -> Literal[
        "monetary_amount",
        "percentage",
        "percentage_range",
        "quantity",
        "measurement",
        "phone_number",
        "date_or_time",
        "account_number",
        "identifier",
        "text",
        "empty",
        "unknown_numeric",
    ]:
        """Classify a table value before applying any monetary normalization.

        Context is deliberately limited to the cell's row and column headers. This
        avoids treating every number in a table on a selected page as money while
        keeping the decision deterministic and auditable.
        """
        text = cell.content.strip().replace("\u00a0", " ")
        if not text:
            return "empty"

        same_row = [
            candidate.content
            for candidate in table_cells
            if candidate.row_index == cell.row_index
        ]
        column_headers = [
            candidate.content
            for candidate in table_cells
            if candidate.column_index == cell.column_index
            and (
                candidate.row_index == 0
                or (candidate.kind or "").casefold() in {"columnheader", "rowheader"}
            )
        ]
        preceding_labels = [
            candidate.content
            for candidate in table_cells
            if candidate.row_index == cell.row_index
            and candidate.column_index == cell.column_index - 1
        ]
        context = " | ".join([*same_row, *column_headers])
        label_context = " | ".join([text, *preceding_labels, *column_headers])
        explicit_currency = bool(
            CURRENCY_CODE_RE.search(text)
            or any(symbol in text for symbol in CURRENCY_SYMBOL_CANDIDATES)
        )
        digit_count = sum(character.isdigit() for character in text)

        if PERCENTAGE_RANGE_RE.fullmatch(text):
            return "percentage_range"
        if PHONE_CONTEXT_RE.search(label_context) and digit_count >= 7:
            return "phone_number"
        if ACCOUNT_CONTEXT_RE.search(label_context) and digit_count >= 4:
            return "account_number"
        if (
            IDENTIFIER_CONTEXT_RE.search(label_context)
            and not DATE_CONTEXT_RE.search(label_context)
            and digit_count
        ):
            return "identifier"
        if DATE_OR_TIME_RE.search(text):
            return "date_or_time"
        if MEASUREMENT_UNIT_RE.search(context) and cls._numeric_candidate(text) is not None:
            return "measurement"
        if QUANTITY_CONTEXT_RE.search(label_context) and cls._numeric_candidate(text) is not None:
            return "quantity"
        if text.endswith("%") and text.count("%") == 1:
            return "percentage"
        if digit_count and (explicit_currency or MONETARY_CONTEXT_RE.search(label_context)):
            return "monetary_amount"
        if cls._numeric_candidate(text) is not None:
            return "unknown_numeric"
        if digit_count >= 7 and re.fullmatch(r"[+()\d\s.-]+", text):
            return "phone_number"
        if digit_count and re.search(r"[A-Za-z]", text) and re.search(r"[-_/]", text):
            return "identifier"
        return "text"

    @staticmethod
    def _currency_contexts(document: CanonicalDocument) -> dict[int, str]:
        contexts: dict[int, str] = {}
        for page in document.pages:
            codes = {
                "CNY" if match.upper() == "RMB" else match.upper()
                for match in CURRENCY_CODE_RE.findall(page.source_text)
            }
            lowered = page.source_text.casefold()
            codes.update(
                code for name, code in CURRENCY_NAME_CONTEXT.items() if name.casefold() in lowered
            )
            if len(codes) == 1:
                contexts[page.page_number] = next(iter(codes))
        return contexts

    @classmethod
    def _infer_decimal_separator(cls, cells: list[TableCell]) -> str | None:
        """Infer a table decimal separator from strong formatting evidence, falling
        back to weak evidence only when nothing conflicts.

        Strong evidence: a cell with both "," and "." (their relative order settles
        which is the decimal point), or a single separator appearing 2+ times with
        valid 3-digit grouping (e.g. "1,234,567" - unambiguously thousands grouping).

        Real "already scaled to thousands" tables (e.g. RMB/USD figures under a
        "thousands" header) commonly have every value under 1,000,000 - a single comma
        group at most - so no cell ever reaches the strong-evidence bar, and every
        such cell was previously left ambiguous. The weak-evidence fallback below
        recognizes a table where every single-comma cell has valid 3-digit grouping
        and none conflicts (an invalidly-grouped single-comma cell, e.g. "12,34",
        disqualifies the whole table) as comma-thousands evidence - extending the same
        grouping assumption the strong tier already makes, just without requiring a
        second comma group to prove it.
        """
        signals: set[str] = set()
        comma_grouped = 0
        comma_ungrouped = 0
        for cell in cells:
            candidate = cls._numeric_candidate(cell.content)
            if candidate is None:
                continue
            unsigned = candidate.lstrip("+-")
            if "," in unsigned and "." in unsigned:
                signals.add("," if unsigned.rfind(",") > unsigned.rfind(".") else ".")
            elif unsigned.count(",") > 1 and cls._valid_grouped_integer(unsigned, ","):
                signals.add(".")
            elif unsigned.count(".") > 1 and cls._valid_grouped_integer(unsigned, "."):
                signals.add(",")
            elif unsigned.count(",") == 1:
                if cls._valid_grouped_integer(unsigned, ","):
                    comma_grouped += 1
                else:
                    comma_ungrouped += 1
        if len(signals) == 1:
            return next(iter(signals))
        if (
            not signals
            and comma_ungrouped == 0
            and comma_grouped >= WEAK_EVIDENCE_MIN_CELLS
        ):
            return "."
        return None

    @staticmethod
    def _valid_grouped_integer(value: str, separator: str) -> bool:
        parts = value.split(separator)
        return bool(parts) and 1 <= len(parts[0]) <= 3 and all(
            len(part) == 3 and part.isdigit() for part in parts[1:]
        )

    @staticmethod
    def _numeric_candidate(raw_text: str) -> str | None:
        text = raw_text.strip().replace("\u00a0", " ")
        text = text.translate(FULLWIDTH_PUNCTUATION_TRANSLATION)
        if not text:
            return None
        has_open_parenthesis = text.startswith("(")
        has_close_parenthesis = text.endswith(")")
        if has_open_parenthesis != has_close_parenthesis:
            return None
        if has_open_parenthesis:
            text = text[1:-1].strip()
        if "%" in text:
            if not text.endswith("%") or text.count("%") != 1:
                return None
            text = text[:-1].strip()
        candidate = text
        for symbol in CURRENCY_SYMBOL_CANDIDATES:
            candidate = candidate.replace(symbol, "")
        candidate = CURRENCY_CODE_RE.sub("", candidate)
        candidate = candidate.strip()
        candidate = re.sub(r"(?<=\d)[ '\u202f](?=\d)", "", candidate)
        if not NUMBER_RE.fullmatch(candidate):
            return None
        return candidate

    @classmethod
    def _normalize(
        cls,
        raw_text: str,
        *,
        semantic_type: Literal[
            "monetary_amount",
            "percentage",
            "percentage_range",
            "quantity",
            "measurement",
            "phone_number",
            "date_or_time",
            "account_number",
            "identifier",
            "text",
            "empty",
            "unknown_numeric",
        ] = "monetary_amount",
        decimal_separator: str | None = None,
        currency_context: str | None = None,
    ) -> NormalizedFinancialValue:
        text = raw_text.strip().replace("\u00a0", " ")
        if not text:
            return NormalizedFinancialValue(
                raw_text=raw_text,
                value_type="empty",
                semantic_type="empty",
                normalization_status="empty",
            )
        if semantic_type not in {"monetary_amount", "percentage"}:
            return NormalizedFinancialValue(
                raw_text=raw_text,
                value_type=(
                    "percentage" if semantic_type == "percentage_range" else "text"
                ),
                semantic_type=semantic_type,
                normalization_status="not_applicable",
            )

        explicit_codes = {
            "CNY" if code.upper() == "RMB" else code.upper()
            for code in CURRENCY_CODE_RE.findall(text)
        }
        present_symbols = {
            symbol for symbol in CURRENCY_SYMBOL_CANDIDATES if symbol in text
        }
        symbol_candidates = {
            code
            for symbol, codes in CURRENCY_SYMBOL_CANDIDATES.items()
            if symbol in present_symbols
            for code in codes
        }
        candidates = explicit_codes or symbol_candidates
        currency: str | None = None
        currency_source: Literal["explicit_code", "symbol", "page_context"] | None = None
        currency_ambiguous = False
        if len(explicit_codes) == 1:
            currency = next(iter(explicit_codes))
            currency_source = "explicit_code"
        elif len(explicit_codes) > 1:
            currency_ambiguous = True
        elif len(symbol_candidates) == 1:
            currency = next(iter(symbol_candidates))
            currency_source = "symbol"
        elif symbol_candidates and currency_context in symbol_candidates:
            currency = currency_context
            currency_source = "page_context"
        elif len(symbol_candidates) > 1:
            currency_ambiguous = True
        if semantic_type == "percentage":
            currency = None
            currency_source = None
            currency_ambiguous = False
            candidates = set()
        currency_fields = {
            "currency": currency,
            "currency_candidates": sorted(candidates),
            "currency_source": currency_source,
            "currency_ambiguous": currency_ambiguous,
            "requires_currency_correction": (
                semantic_type == "monetary_amount" and currency_ambiguous
            ),
        }
        malformed_numeric_wrapper = any(character.isdigit() for character in text) and (
            text.startswith("(") != text.endswith(")")
            or ("%" in text and (not text.endswith("%") or text.count("%") != 1))
        )
        if (
            malformed_numeric_wrapper
            or len(explicit_codes) > 1
            or len(present_symbols) > 1
        ):
            return NormalizedFinancialValue(
                raw_text=raw_text,
                semantic_type=semantic_type,
                **currency_fields,
                ambiguous=True,
                requires_normalized_correction=semantic_type == "monetary_amount",
                review_reason_code=(
                    "malformed_monetary_value"
                    if semantic_type == "monetary_amount"
                    else None
                ),
                normalization_status="ambiguous",
            )
        parenthesized_negative = text.startswith("(") and text.endswith(")")
        percentage = text.endswith("%")
        if semantic_type == "monetary_amount":
            magnitude_source = text[1:-1].strip() if parenthesized_negative else text
            magnitude_source = magnitude_source.translate(FULLWIDTH_PUNCTUATION_TRANSLATION)
            for name in sorted(CURRENCY_NAME_CONTEXT, key=len, reverse=True):
                if magnitude_source.startswith(name):
                    magnitude_source = magnitude_source[len(name) :].strip()
                    break
            for symbol in CURRENCY_SYMBOL_CANDIDATES:
                magnitude_source = magnitude_source.replace(symbol, "")
            magnitude_source = CURRENCY_CODE_RE.sub("", magnitude_source).strip()
            magnitude_match = CJK_MAGNITUDE_RE.match(magnitude_source)
            if magnitude_match:
                # A distinct code path from the comma/dot ambiguity logic below - it
                # never runs on this input. Always flagged for reviewer sign-off in
                # this first release (requires_normalized_correction=True) even though
                # a value was computed, matching the app's recall-first philosophy for
                # a newer, less-proven parsing feature.
                magnitude_value = Decimal(
                    magnitude_match.group("number")
                ) * CJK_MAGNITUDE_MULTIPLIERS[magnitude_match.group("magnitude")]
                if parenthesized_negative:
                    magnitude_value = -magnitude_value
                return NormalizedFinancialValue(
                    raw_text=raw_text,
                    semantic_type=semantic_type,
                    **currency_fields,
                    normalized_value=magnitude_value,
                    negative=parenthesized_negative,
                    ambiguous=True,
                    requires_normalized_correction=True,
                    review_reason_code="cjk_magnitude_suffix",
                    normalization_status="parsed",
                )
        candidate = cls._numeric_candidate(raw_text)
        if candidate is None:
            return NormalizedFinancialValue(
                raw_text=raw_text,
                semantic_type=semantic_type,
                **currency_fields,
                ambiguous=semantic_type == "monetary_amount",
                requires_normalized_correction=semantic_type == "monetary_amount",
                review_reason_code=(
                    "malformed_monetary_value"
                    if semantic_type == "monetary_amount"
                    else None
                ),
                normalization_status=(
                    "ambiguous" if semantic_type == "monetary_amount" else "unparsed"
                ),
            )

        sign = ""
        unsigned = candidate
        if unsigned[0] in "+-":
            sign, unsigned = unsigned[0], unsigned[1:]
        if not unsigned or not all(
            character.isdigit() or character in ",." for character in unsigned
        ):
            return NormalizedFinancialValue(
                raw_text=raw_text,
                semantic_type=semantic_type,
                **currency_fields,
                normalization_status="unparsed",
            )

        canonical: str | None = None
        if "," in unsigned and "." in unsigned:
            inferred_decimal = "," if unsigned.rfind(",") > unsigned.rfind(".") else "."
            grouping = "." if inferred_decimal == "," else ","
            if unsigned.count(inferred_decimal) == 1:
                integer_part, fractional_part = unsigned.rsplit(inferred_decimal, 1)
                grouping_valid = grouping not in integer_part or cls._valid_grouped_integer(
                    integer_part, grouping
                )
                if grouping_valid and fractional_part.isdigit() and fractional_part:
                    canonical = integer_part.replace(grouping, "") + "." + fractional_part
        elif "," in unsigned or "." in unsigned:
            separator = "," if "," in unsigned else "."
            count = unsigned.count(separator)
            if count > 1:
                if cls._valid_grouped_integer(unsigned, separator):
                    canonical = unsigned.replace(separator, "")
            elif decimal_separator is not None:
                integer_part, fractional_part = unsigned.split(separator, 1)
                if separator == decimal_separator and fractional_part:
                    canonical = integer_part + "." + fractional_part
                elif cls._valid_grouped_integer(unsigned, separator):
                    canonical = unsigned.replace(separator, "")
            if canonical is None:
                return NormalizedFinancialValue(
                    raw_text=raw_text,
                    semantic_type=semantic_type,
                    **currency_fields,
                    ambiguous=True,
                    requires_normalized_correction=semantic_type == "monetary_amount",
                    review_reason_code=(
                        "ambiguous_monetary_separator"
                        if semantic_type == "monetary_amount"
                        else None
                    ),
                    normalization_status="ambiguous",
                )
        elif unsigned.isdigit():
            canonical = unsigned

        if canonical is None:
            return NormalizedFinancialValue(
                raw_text=raw_text,
                semantic_type=semantic_type,
                **currency_fields,
                ambiguous=True,
                requires_normalized_correction=semantic_type == "monetary_amount",
                review_reason_code=(
                    "ambiguous_monetary_separator"
                    if semantic_type == "monetary_amount"
                    else None
                ),
                normalization_status="ambiguous",
            )

        candidate_value = sign + canonical
        try:
            value = Decimal(candidate_value)
        except InvalidOperation:
            return NormalizedFinancialValue(
                raw_text=raw_text,
                semantic_type=semantic_type,
                **currency_fields,
                ambiguous=True,
                requires_normalized_correction=semantic_type == "monetary_amount",
                review_reason_code=(
                    "malformed_monetary_value"
                    if semantic_type == "monetary_amount"
                    else None
                ),
                normalization_status="ambiguous",
            )
        if parenthesized_negative:
            value = -abs(value)
        return NormalizedFinancialValue(
            raw_text=raw_text,
            normalized_value=value,
            value_type="percentage" if percentage else "amount",
            semantic_type=semantic_type,
            **currency_fields,
            negative=value < 0,
            ambiguous=False,
            review_reason_code=(
                "ambiguous_currency"
                if semantic_type == "monetary_amount" and currency_ambiguous
                else None
            ),
            normalization_status="parsed",
        )

    @staticmethod
    def _validate(
        classification: FinancialClassificationResult,
        tables: list[FinancialTable],
        content_items: list[FinancialContentItem],
    ) -> list[FinancialValidationIssue]:
        issues: list[FinancialValidationIssue] = []

        def add(
            code: str,
            message: str,
            *,
            page: int | None = None,
            table: str | None = None,
            cell: str | None = None,
            severity: str = "warning",
        ) -> None:
            digest = hashlib.sha256(f"{code}:{page}:{table}:{cell}".encode()).hexdigest()[:16]
            issues.append(
                FinancialValidationIssue(
                    issue_id=f"fv-{digest}",
                    severity="error" if severity == "error" else "warning",
                    code=code,
                    message=message,
                    page_number=page,
                    table_id=table,
                    cell_id=cell,
                )
            )

        financial_content_pages = {
            page
            for item in content_items
            for page in item.source_pages
        }
        for page in classification.pages:
            if page.selected and page.page_number not in financial_content_pages:
                add(
                    "selected_page_without_financial_content",
                    "The selected page contains no extracted financial content.",
                    page=page.page_number,
                )
        for table in tables:
            if table.integrity_status == "reconciled":
                add(
                    "table_structure_reconstructed",
                    "Aligned source blocks were joined to a provider table; a reviewer must "
                    "accept or reject the reconstructed structure.",
                    page=table.source_pages[0],
                    table=table.table_id,
                )
            if not table.complete:
                add(
                    "incomplete_multi_page_table",
                    "The provider marked the table as spanning pages for which no cells "
                    "were returned.",
                    page=table.source_pages[0],
                    table=table.table_id,
                    severity="error",
                )
            if table.classification_override_pages:
                add(
                    "table_crosses_selection_boundary",
                    "The complete provider table was retained across a page-classification "
                    "boundary and requires review.",
                    page=table.classification_override_pages[0],
                    table=table.table_id,
                )
            coordinates: set[tuple[int, int]] = set()
            for cell in table.cells:
                coordinate = (cell.row_index, cell.column_index)
                if coordinate in coordinates:
                    add(
                        "duplicate_cell_coordinate",
                        "Two cells share the same table coordinate.",
                        page=cell.source_page,
                        table=table.table_id,
                        cell=cell.cell_id,
                        severity="error",
                    )
                coordinates.add(coordinate)
                if cell.value.requires_normalized_correction:
                    add(
                        "ambiguous_monetary_format",
                        "The monetary value cannot be normalized safely and needs review.",
                        page=cell.source_page,
                        table=table.table_id,
                        cell=cell.cell_id,
                    )
                if cell.value.requires_currency_correction:
                    add(
                        "ambiguous_currency_symbol",
                        "The currency symbol maps to multiple currencies and needs review.",
                        page=cell.source_page,
                        table=table.table_id,
                        cell=cell.cell_id,
                    )
            if len(table.currencies) > 1:
                add(
                    "mixed_table_currencies",
                    "The table contains more than one explicit currency.",
                    page=table.source_pages[0],
                    table=table.table_id,
                )
        return issues
